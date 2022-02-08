import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Planilha1']

#Buscando valores
print(planilha1['b4'].value)
print('-----------------')

for campo in planilha1['b']:
    print(campo.value)

print('-----------------')

for linha in planilha1['a1:c2']:
    for coluna in linha:
        print(coluna.value)

#Manipulando valores (em outra planilha)
planilha1['B3'].value = 1000

#começa na linha 5 e para na 15
for linha in range(5, 16):
    numero_pedido = linha - 1 
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    #uniform -> numero de 10 a 100
    #round -> número de casas decimais
    preco = round(uniform(10, 100), 2) 
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')

#Criando uma planilha do zero

new_planilha = openpyxl.Workbook()
new_planilha.create_sheet('Planilha1', 0) #nome e índice que será criada

new_planilha1 = new_planilha['Planilha1']

new_planilha1['a1'].value = 'A'
new_planilha1.cell(1, 2).value = 'B'

new_planilha.save('planilha_teste.xlsx')